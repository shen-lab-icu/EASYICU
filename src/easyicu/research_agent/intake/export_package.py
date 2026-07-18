"""Manifest-authoritative intake for EasyICU native and legacy exports.

The Web application writes ``_manifest.json`` while older headless exporters
write ``easyicu_export_manifest.json``.  This leaf adapter accepts both, with
the native marker taking precedence, but never discovers data by globbing the
directory: only files named by the selected manifest can become data
authority.  It intentionally does not import the Web layer.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import tempfile
from contextlib import contextmanager
from dataclasses import dataclass
from pathlib import Path
from types import MappingProxyType
from typing import Any, Dict, Iterator, Mapping, Optional, Sequence, Tuple, Union

import pandas as pd
import pyarrow.parquet as pq

from easyicu.database_config import ID_COLUMNS as HOST_DATABASE_ID_COLUMNS
from easyicu.database_config import END_TIME_COLUMNS as HOST_DATABASE_END_TIME_COLUMNS
from easyicu.database_config import (
    START_TIME_COLUMNS as HOST_DATABASE_START_TIME_COLUMNS,
)
from easyicu.database_config import TIME_COLUMNS as HOST_DATABASE_TIME_COLUMNS

ID_COL = "stay_id"
PRIMARY_ID_COLUMNS = MappingProxyType(
    {
        **{
            database: columns[0]
            for database, columns in HOST_DATABASE_ID_COLUMNS.items()
        },
        "mimic": "icustay_id",
        "mimiciii": "icustay_id",
    }
)
IDENTIFIER_COLUMNS = frozenset(
    {
        ID_COL,
        *(
            column
            for columns in HOST_DATABASE_ID_COLUMNS.values()
            for column in columns
        ),
        *PRIMARY_ID_COLUMNS.values(),
        "subject_id",
        "hadm_id",
    }
)
TIME_COLUMNS = frozenset(
    {
        "charttime",
        "time",
        *HOST_DATABASE_TIME_COLUMNS.values(),
        *HOST_DATABASE_START_TIME_COLUMNS.values(),
        *HOST_DATABASE_END_TIME_COLUMNS.values(),
        "starttime",
        "start",
        "measuredat_minutes",
        "givenat",
        "infusionoffset",
        "labresultoffset",
    }
)
NATIVE_MANIFEST = "_manifest.json"
LEGACY_MANIFEST = "easyicu_export_manifest.json"
FEATURE_DEFINITION_SCHEMA = "easyicu_feature_definitions_v1"
_MAX_JSON_BYTES = 32 * 1024 * 1024
_FORMAT_BY_SUFFIX = {
    ".csv": "csv",
    ".parquet": "parquet",
    ".xlsx": "excel",
}
_AUTHORIZED_COMPANION_SUFFIXES = (
    "_n",
    "_count",
    "_measured",
    "_first_time",
    "_last_time",
    "_max",
    "_min",
    "_mean",
    "_median",
    "_first",
    "_last",
)
_FORMAT_ALIASES = {
    "csv": "csv",
    "parquet": "parquet",
    "excel": "excel",
    "xlsx": "excel",
}


class ExportPackageError(ValueError):
    """Raised when an export package cannot establish trusted authority."""

    def __init__(
        self,
        message: str,
        *,
        code: str = "export_package_invalid",
        manifest_path: Optional[Path] = None,
        member: Optional[str] = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.manifest_path = manifest_path
        self.member = member


@dataclass(frozen=True, slots=True)
class FileIdentity:
    """Cheap mutation fence for one inspected package member."""

    device: int
    inode: int
    size: int
    mtime_ns: int
    sha256: str

    @classmethod
    def from_handle(cls, handle: Any) -> "FileIdentity":
        stat = os.fstat(handle.fileno())
        digest = hashlib.sha256()
        original_position = handle.tell()
        try:
            handle.seek(0)
            while chunk := handle.read(1024 * 1024):
                digest.update(chunk)
        finally:
            handle.seek(original_position)
        after = os.fstat(handle.fileno())

        def coordinates(value: os.stat_result) -> tuple[int, int, int, int]:
            return (
                int(value.st_dev),
                int(value.st_ino),
                int(value.st_size),
                int(value.st_mtime_ns),
            )

        if coordinates(after) != coordinates(stat):
            raise ExportPackageError(
                "manifest-listed file changed while hashing",
                code="manifest_file_mutated",
            )
        return cls(
            device=int(stat.st_dev),
            inode=int(stat.st_ino),
            size=int(stat.st_size),
            mtime_ns=int(stat.st_mtime_ns),
            sha256=digest.hexdigest(),
        )

    @classmethod
    def from_path(cls, path: Path) -> "FileIdentity":
        with path.open("rb") as handle:
            identity = cls.from_handle(handle)
        stat = path.stat()
        if (
            int(stat.st_dev),
            int(stat.st_ino),
            int(stat.st_size),
            int(stat.st_mtime_ns),
        ) != (
            identity.device,
            identity.inode,
            identity.size,
            identity.mtime_ns,
        ):
            raise ExportPackageError(
                f"manifest-listed path changed while hashing: {path}",
                code="manifest_file_mutated",
                member=path.name,
            )
        return identity


def _require_identity(path: Path, expected: FileIdentity) -> None:
    """Fail closed when a package member changes after verification."""

    try:
        current = FileIdentity.from_path(path)
    except OSError as exc:
        raise ExportPackageError(
            f"manifest-listed file disappeared after verification: {path}",
            code="manifest_file_mutated",
            member=path.name,
        ) from exc
    if current != expected:
        raise ExportPackageError(
            f"manifest-listed file changed after verification: {path}",
            code="manifest_file_mutated",
            member=path.name,
        )


@contextmanager
def _verified_file_snapshot(
    path: Path,
    *,
    expected: Optional[FileIdentity] = None,
    member: Optional[str] = None,
) -> Iterator[tuple[Any, FileIdentity]]:
    """Yield a host-owned snapshot containing exactly the verified bytes."""

    label = member or path.name
    try:
        with (
            path.open("rb") as source,
            tempfile.TemporaryFile(mode="w+b") as snapshot,
        ):
            before = os.fstat(source.fileno())
            before_coordinates = (
                int(before.st_dev),
                int(before.st_ino),
                int(before.st_size),
                int(before.st_mtime_ns),
            )
            if expected is not None and before_coordinates != (
                expected.device,
                expected.inode,
                expected.size,
                expected.mtime_ns,
            ):
                raise ExportPackageError(
                    f"manifest-listed file identity changed before snapshot: {label}",
                    code="manifest_file_mutated",
                    member=label,
                )

            digest = hashlib.sha256()
            while chunk := source.read(1024 * 1024):
                digest.update(chunk)
                snapshot.write(chunk)
            after = os.fstat(source.fileno())
            observed = FileIdentity(
                device=int(after.st_dev),
                inode=int(after.st_ino),
                size=int(after.st_size),
                mtime_ns=int(after.st_mtime_ns),
                sha256=digest.hexdigest(),
            )
            if before_coordinates != (
                observed.device,
                observed.inode,
                observed.size,
                observed.mtime_ns,
            ) or (expected is not None and observed != expected):
                raise ExportPackageError(
                    f"manifest-listed file changed while creating snapshot: {label}",
                    code="manifest_file_mutated",
                    member=label,
                )

            snapshot.flush()
            snapshot.seek(0)
            yield snapshot, observed
        _require_identity(path, observed)
    except ExportPackageError:
        raise
    except OSError as exc:
        raise ExportPackageError(
            f"cannot snapshot manifest-authorized file: {path}",
            code="manifest_file_mutated",
            member=label,
        ) from exc


@dataclass(frozen=True, slots=True)
class ExportPhysicalFile:
    """One manifest-listed physical data file after schema inspection."""

    path: Path
    relative_path: str
    module: str
    file_format: str
    rows: Optional[int]
    columns: Tuple[str, ...]
    declared_concepts: Tuple[str, ...]
    id_column: str
    time_column: Optional[str]
    time_columns: Tuple[str, ...]
    identity: FileIdentity
    excel_sheet: Optional[str] = None


@dataclass(frozen=True, slots=True)
class ExportPackage:
    """Verified immutable view of one EasyICU export directory."""

    root: Path
    manifest_path: Path
    manifest_kind: str
    manifest_sha256: str
    authority_sha256: str
    database: str
    export_format: str
    files: Tuple[ExportPhysicalFile, ...]
    concept_index: Mapping[str, Mapping[str, object]]
    missing_selected_concepts: Tuple[str, ...]
    feature_definitions: Mapping[str, Mapping[str, object]]
    feature_definitions_sha256: Optional[str]

    def index_dict(self) -> Dict[str, Dict[str, object]]:
        """Return the compatibility index shape used by existing callers."""

        return {concept: dict(info) for concept, info in self.concept_index.items()}


def _read_json(path: Path, *, label: str) -> tuple[dict[str, Any], bytes]:
    try:
        with _verified_file_snapshot(path, member=path.name) as (snapshot, _identity):
            raw = snapshot.read(_MAX_JSON_BYTES + 1)
    except OSError as exc:
        raise ExportPackageError(
            f"cannot read {label}: {path}", code=f"{label.replace(' ', '_')}_missing"
        ) from exc
    if len(raw) > _MAX_JSON_BYTES:
        raise ExportPackageError(f"{label} exceeds {_MAX_JSON_BYTES} bytes: {path}")
    try:
        payload = json.loads(raw.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ExportPackageError(
            f"invalid JSON in {label}: {path}",
            code=f"{label.replace(' ', '_')}_json_invalid",
        ) from exc
    if not isinstance(payload, dict):
        raise ExportPackageError(f"{label} must contain a JSON object: {path}")
    return payload, raw


def _safe_manifest_file(root: Path, value: object, *, label: str) -> Path:
    if not isinstance(value, str) or not value.strip():
        raise ExportPackageError(f"{label} must name a non-empty relative file")
    relative = Path(value)
    if (
        relative.is_absolute()
        or ".." in relative.parts
        or "\\" in value
        or "\x00" in value
    ):
        raise ExportPackageError(
            f"{label} escapes the export root: {value!r}",
            code="manifest_path_escape",
            member=value,
        )
    candidate = root.joinpath(relative)
    cursor = root
    for part in relative.parts:
        cursor = cursor / part
        if cursor.is_symlink():
            raise ExportPackageError(
                f"{label} must not traverse a symlink: {value!r}",
                code="manifest_file_symlink",
                member=value,
            )
    try:
        resolved = candidate.resolve(strict=True)
    except OSError as exc:
        raise ExportPackageError(
            f"manifest-listed file is missing: {value!r}",
            code="manifest_file_missing",
            member=value,
        ) from exc
    resolved_root = root.resolve(strict=True)
    if not resolved.is_relative_to(resolved_root) or not resolved.is_file():
        raise ExportPackageError(
            f"{label} is outside the export root: {value!r}",
            code="manifest_path_escape",
            member=value,
        )
    return resolved


def _normalize_format(value: object) -> Optional[str]:
    if value is None or str(value).strip() == "":
        return None
    return _FORMAT_ALIASES.get(str(value).strip().lower())


def _manifest_file_entries(
    payload: Mapping[str, Any], *, manifest_kind: str
) -> list[dict[str, Any]]:
    if manifest_kind == "native":
        raw_files = payload.get("files")
        if not isinstance(raw_files, list) or not raw_files:
            raise ExportPackageError(
                "native manifest must contain a non-empty files list"
            )
        entries = []
        for raw in raw_files:
            if not isinstance(raw, Mapping):
                raise ExportPackageError(
                    "native manifest files entries must be objects"
                )
            entries.append(dict(raw))
        return entries

    inventories: list[list[dict[str, Any]]] = []
    modules = payload.get("modules")
    if isinstance(modules, list):
        module_entries: list[dict[str, Any]] = []
        for raw in modules:
            if isinstance(raw, Mapping) and raw.get("file"):
                module_entries.append(dict(raw))
        if module_entries:
            inventories.append(module_entries)
    exported_files = payload.get("exported_files")
    if isinstance(exported_files, list):
        inventories.append([{"file": value} for value in exported_files])
    raw_files = payload.get("files")
    if isinstance(raw_files, list):
        inventories.append(
            [
                dict(raw) if isinstance(raw, Mapping) else {"file": raw}
                for raw in raw_files
            ]
        )

    inventories = [
        [entry for entry in inventory if str(entry.get("file") or "")]
        for inventory in inventories
    ]
    inventories = [inventory for inventory in inventories if inventory]
    if not inventories:
        raise ExportPackageError(
            "legacy manifest does not list physical files; directory globbing is not authority",
            code="manifest_inventory_missing",
        )

    normalized_inventories: list[tuple[list[dict[str, Any]], set[str]]] = []
    for inventory in inventories:
        names = [Path(str(entry["file"])).as_posix() for entry in inventory]
        if len(set(names)) != len(names):
            raise ExportPackageError(
                "legacy manifest lists the same file more than once",
                code="manifest_duplicate_file",
            )
        normalized_inventories.append((inventory, set(names)))
    canonical_set = normalized_inventories[0][1]
    if any(names != canonical_set for _, names in normalized_inventories[1:]):
        raise ExportPackageError(
            "legacy manifest file inventories disagree",
            code="manifest_inventory_mismatch",
        )

    # Prefer the richest compatible inventory while treating the other legacy
    # fields as redundant cross-checks, not duplicate file declarations.
    entries = max(
        (inventory for inventory, _ in normalized_inventories),
        key=lambda values: sum(len(value) for value in values),
    )

    deduped: list[dict[str, Any]] = []
    seen: set[str] = set()
    for entry in entries:
        file_name = str(entry.get("file") or "")
        if not file_name:
            continue
        if file_name in seen:
            raise ExportPackageError(
                f"manifest lists the same file more than once: {file_name!r}",
                code="manifest_duplicate_file",
                member=file_name,
            )
        seen.add(file_name)
        deduped.append(entry)
    return deduped


def _validated_columns(values: Sequence[object], *, path: Path) -> Tuple[str, ...]:
    columns = tuple("" if value is None else str(value) for value in values)
    if not columns or any(not value for value in columns):
        raise ExportPackageError(
            f"export file has an empty header: {path}", code="file_schema_invalid"
        )
    if len(set(columns)) != len(columns):
        raise ExportPackageError(
            f"export file has duplicate columns: {path}", code="file_schema_invalid"
        )
    return columns


def _schema_for_file(
    path: Path, file_format: str
) -> tuple[Tuple[str, ...], int | None, Optional[str], FileIdentity]:
    try:
        with _verified_file_snapshot(path, member=path.name) as (
            snapshot,
            identity,
        ):
            if file_format == "parquet":
                parquet_file = pq.ParquetFile(snapshot)
                columns = _validated_columns(parquet_file.schema.names, path=path)
                rows = int(parquet_file.metadata.num_rows)
                excel_sheet = None
            elif file_format == "csv":
                import io

                text = io.TextIOWrapper(
                    snapshot, encoding="utf-8-sig", newline="", write_through=True
                )
                reader = csv.reader(text)
                columns = _validated_columns(next(reader, ()), path=path)
                rows = sum(1 for _ in reader)
                excel_sheet = None
                text.detach()
            elif file_format == "excel":
                from openpyxl import load_workbook

                workbook = load_workbook(snapshot, read_only=True, data_only=True)
                try:
                    nonempty: list[tuple[str, Tuple[str, ...]]] = []
                    for sheet in workbook.worksheets:
                        header = next(
                            sheet.iter_rows(min_row=1, max_row=1, values_only=True),
                            (),
                        )
                        if any(
                            value is not None and str(value) != "" for value in header
                        ):
                            nonempty.append(
                                (sheet.title, _validated_columns(header, path=path))
                            )
                    if len(nonempty) != 1:
                        raise ExportPackageError(
                            "Excel export must contain exactly one non-empty sheet: "
                            f"{path}",
                            code="file_schema_invalid",
                        )
                    excel_sheet, columns = nonempty[0]
                    rows = max(0, int(workbook[excel_sheet].max_row) - 1)
                finally:
                    workbook.close()
            else:  # pragma: no cover - guarded by caller
                raise ExportPackageError(f"unsupported export format: {file_format}")
    except ExportPackageError:
        raise
    except Exception as exc:
        raise ExportPackageError(f"cannot inspect export file: {path}") from exc
    return columns, rows, excel_sheet, identity


def _declared_concepts(entry: Mapping[str, Any]) -> Tuple[str, ...]:
    values = entry.get("concept_ids")
    if not isinstance(values, list):
        return ()
    if not all(isinstance(value, str) for value in values):
        raise ExportPackageError(
            "manifest concept_ids must contain strings",
            code="manifest_concept_ids_invalid",
        )
    concepts = tuple(value.strip() for value in values)
    if any(not value for value in concepts) or len(set(concepts)) != len(concepts):
        raise ExportPackageError(
            "manifest concept_ids must be unique and non-empty",
            code="manifest_concept_ids_invalid",
        )
    return concepts


def _load_feature_definitions(
    *,
    root: Path,
    manifest: Mapping[str, Any],
    selected_concepts: set[str],
    selected_by_module: Mapping[str, set[str]],
    files_by_module: Mapping[str, set[str]],
    authorized_files: set[str],
    database: str,
) -> tuple[Mapping[str, Mapping[str, object]], Optional[str]]:
    descriptor = manifest.get("feature_definitions")
    if descriptor is None:
        return MappingProxyType({}), None
    if not isinstance(descriptor, Mapping):
        raise ExportPackageError("feature_definitions manifest entry must be an object")
    included = descriptor.get("included")
    if not isinstance(included, bool):
        raise ExportPackageError(
            "feature_definitions.included must be boolean",
            code="feature_definition_inventory_invalid",
        )
    if not included:
        return MappingProxyType({}), None

    schema_version = str(descriptor.get("schema_version") or "")
    if schema_version != FEATURE_DEFINITION_SCHEMA:
        raise ExportPackageError(
            f"unsupported feature definition schema: {schema_version!r}"
        )
    definition_files = descriptor.get("files")
    if not isinstance(definition_files, list) or not definition_files:
        raise ExportPackageError("included feature definitions must list files")
    json_names: list[str] = []
    seen_files: set[str] = set()
    seen_kinds: set[str] = set()
    descriptor_count = descriptor.get("record_count")
    if (
        not isinstance(descriptor_count, int)
        or isinstance(descriptor_count, bool)
        or descriptor_count < 0
    ):
        raise ExportPackageError(
            "feature definition record_count must be a non-negative integer",
            code="feature_definition_count_mismatch",
        )
    for raw in definition_files:
        if not isinstance(raw, Mapping):
            raise ExportPackageError("feature definition file entries must be objects")
        raw_file_name = str(raw.get("file") or "")
        file_name = Path(raw_file_name).as_posix() if raw_file_name else ""
        if not file_name or file_name in seen_files:
            raise ExportPackageError(
                "feature definition files must be unique and non-empty",
                code="feature_definition_inventory_invalid",
            )
        seen_files.add(file_name)
        kind = str(raw.get("kind") or "")
        if not kind or kind in seen_kinds:
            raise ExportPackageError(
                "feature definition file kinds must be unique and non-empty",
                code="feature_definition_inventory_invalid",
            )
        seen_kinds.add(kind)
        _safe_manifest_file(root, file_name, label="feature definition file")
        raw_records = raw.get("records")
        if raw_records is not None and (
            not isinstance(raw_records, int)
            or isinstance(raw_records, bool)
            or raw_records != descriptor_count
        ):
            raise ExportPackageError(
                "feature definition file count does not match descriptor",
                code="feature_definition_count_mismatch",
            )
        if kind == "feature_definitions":
            json_names.append(file_name)
    if len(json_names) != 1:
        raise ExportPackageError(
            "included feature definitions require exactly one JSON authority file",
            code="feature_definition_authority_invalid",
        )
    json_name = json_names[0]
    definition_path = _safe_manifest_file(
        root, json_name, label="feature definition file"
    )
    payload, raw_bytes = _read_json(definition_path, label="feature definitions")
    if payload.get("schema_version") != FEATURE_DEFINITION_SCHEMA:
        raise ExportPackageError(
            "feature definition payload schema does not match manifest"
        )
    payload_database = str(payload.get("database") or "")
    if database and payload_database != database:
        raise ExportPackageError(
            "feature definition database does not match export manifest",
            code="feature_definition_database_mismatch",
        )
    records = payload.get("records")
    if not isinstance(records, list) or not all(
        isinstance(record, Mapping) for record in records
    ):
        raise ExportPackageError("feature definition records must be a list of objects")
    declared_count = descriptor_count
    payload_count = payload.get("record_count")
    if (
        not isinstance(payload_count, int)
        or isinstance(payload_count, bool)
        or declared_count != len(records)
        or payload_count != len(records)
    ):
        raise ExportPackageError(
            "feature definition record count does not match payload"
        )

    by_concept: dict[str, Mapping[str, object]] = {}
    for raw_record in records:
        record = dict(raw_record)
        concept = str(record.get("concept_id") or "")
        if not concept or concept in by_concept:
            raise ExportPackageError(
                "feature definition concepts must be unique and non-empty"
            )
        record_database = str(record.get("database") or "")
        if database and record_database != database:
            raise ExportPackageError(
                f"feature definition database mismatch for {concept!r}",
                code="feature_definition_database_mismatch",
            )
        module = str(record.get("module") or "")
        if module not in files_by_module:
            raise ExportPackageError(
                f"feature definition module is not exported: {module!r}",
                code="feature_definition_module_mismatch",
            )
        if selected_by_module and concept not in selected_by_module.get(module, set()):
            raise ExportPackageError(
                f"feature definition concept is not selected in module {module!r}",
                code="feature_definition_module_mismatch",
            )
        source = record.get("source")
        if not isinstance(source, Mapping):
            raise ExportPackageError(
                f"feature definition source is missing for {concept!r}",
                code="feature_definition_source_invalid",
            )
        export_files = source.get("export_files")
        if not isinstance(export_files, list) or not export_files:
            raise ExportPackageError(
                f"feature definition export_files are missing for {concept!r}",
                code="feature_definition_source_invalid",
            )
        if not all(isinstance(value, str) for value in export_files):
            raise ExportPackageError(
                f"feature definition export_files are invalid for {concept!r}",
                code="feature_definition_source_invalid",
            )
        raw_export_files = [value.strip() for value in export_files]
        if any(not value for value in raw_export_files) or len(
            set(raw_export_files)
        ) != len(raw_export_files):
            raise ExportPackageError(
                f"feature definition export_files are invalid for {concept!r}",
                code="feature_definition_source_invalid",
            )
        normalized_files = {Path(value).as_posix() for value in raw_export_files}
        if not normalized_files.issubset(
            authorized_files
        ) or not normalized_files.issubset(files_by_module[module]):
            raise ExportPackageError(
                f"feature definition source files are not authorized for {concept!r}",
                code="feature_definition_source_mismatch",
            )
        by_concept[concept] = MappingProxyType(record)
    if selected_concepts and set(by_concept) != selected_concepts:
        raise ExportPackageError(
            "feature definition concepts do not match the manifest concept selection"
        )
    return (
        MappingProxyType(by_concept),
        hashlib.sha256(raw_bytes).hexdigest(),
    )


def is_export_package(path: Union[str, Path]) -> bool:
    """Return whether ``path`` carries a native or legacy root marker."""

    root = Path(path).expanduser()
    return root.is_dir() and (
        _marker_present(root.joinpath(NATIVE_MANIFEST))
        or _marker_present(root.joinpath(LEGACY_MANIFEST))
    )


def _marker_present(path: Path) -> bool:
    """Return marker presence without following a dangling symlink."""

    return path.exists() or path.is_symlink()


def _select_manifest(root: Path) -> tuple[Path, str]:
    native_path = root / NATIVE_MANIFEST
    legacy_path = root / LEGACY_MANIFEST
    if _marker_present(native_path):
        if native_path.is_symlink() or not native_path.is_file():
            raise ExportPackageError(
                f"native export marker is not a regular file: {native_path}",
                code="manifest_marker_invalid",
                manifest_path=native_path,
            )
        return native_path, "native"
    if _marker_present(legacy_path):
        if legacy_path.is_symlink() or not legacy_path.is_file():
            raise ExportPackageError(
                f"legacy export marker is not a regular file: {legacy_path}",
                code="manifest_marker_invalid",
                manifest_path=legacy_path,
            )
        return legacy_path, "legacy"
    raise ExportPackageError(
        f"export package lacks {NATIVE_MANIFEST!r} or {LEGACY_MANIFEST!r}: {root}",
        code="manifest_marker_missing",
    )


def _native_selection(
    manifest: Mapping[str, Any], entries: Sequence[Mapping[str, Any]]
) -> tuple[set[str], Mapping[str, set[str]]]:
    database = str(manifest.get("database") or "").strip()
    if not database:
        raise ExportPackageError(
            "native manifest database must be non-empty",
            code="manifest_database_missing",
        )
    declared_by_module: dict[str, set[str]] = {}
    for entry in entries:
        module = str(entry.get("module") or "").strip()
        if not module:
            raise ExportPackageError(
                "native manifest files require a module",
                code="manifest_module_missing",
            )
        concepts = _declared_concepts(entry)
        if not concepts:
            raise ExportPackageError(
                "native manifest files require concept_ids",
                code="manifest_concept_ids_invalid",
            )
        raw_count = entry.get("concepts")
        if not isinstance(raw_count, int) or isinstance(raw_count, bool):
            raise ExportPackageError(
                "native manifest concept count must be an integer",
                code="manifest_concept_count_mismatch",
            )
        concept_count = raw_count
        if concept_count != len(concepts):
            raise ExportPackageError(
                "native manifest concept count does not match concept_ids",
                code="manifest_concept_count_mismatch",
            )
        declared_by_module.setdefault(module, set()).update(concepts)

    selection = manifest.get("concept_selection")
    modules = selection.get("modules") if isinstance(selection, Mapping) else None
    if not isinstance(modules, Mapping) or not modules:
        raise ExportPackageError(
            "native manifest concept_selection.modules must be non-empty",
            code="manifest_concept_selection_invalid",
        )
    selected_by_module: dict[str, set[str]] = {}
    for raw_module, raw_concepts in modules.items():
        module = str(raw_module).strip()
        if not module or not isinstance(raw_concepts, list):
            raise ExportPackageError(
                "native concept selection modules must contain concept lists",
                code="manifest_concept_selection_invalid",
            )
        if not all(isinstance(value, str) for value in raw_concepts):
            raise ExportPackageError(
                "native concept selection must contain strings",
                code="manifest_concept_selection_invalid",
            )
        concepts = [value.strip() for value in raw_concepts]
        if any(not value for value in concepts) or len(set(concepts)) != len(concepts):
            raise ExportPackageError(
                "native concept selection must be unique and non-empty",
                code="manifest_concept_selection_invalid",
            )
        selected_by_module[module] = set(concepts)
    if selected_by_module != declared_by_module:
        raise ExportPackageError(
            "native concept selection does not match file declarations",
            code="manifest_concept_selection_mismatch",
        )
    return set().union(*selected_by_module.values()), MappingProxyType(
        selected_by_module
    )


def _identity_column(columns: Sequence[str], database: str, *, path: Path) -> str:
    if ID_COL in columns:
        return ID_COL
    preferred = PRIMARY_ID_COLUMNS.get(database.strip().lower())
    if preferred and preferred in columns:
        return preferred
    candidates = [column for column in columns if column in IDENTIFIER_COLUMNS]
    if len(candidates) != 1:
        raise ExportPackageError(
            f"export file requires one unambiguous ICU-stay identifier: {path}",
            code="file_identity_column_invalid",
            member=path.name,
        )
    return candidates[0]


def _time_columns(columns: Sequence[str], database: str) -> Tuple[str, ...]:
    preferred = (
        "charttime",
        "time",
        HOST_DATABASE_TIME_COLUMNS.get(database.strip().lower(), ""),
        HOST_DATABASE_START_TIME_COLUMNS.get(database.strip().lower(), ""),
        HOST_DATABASE_END_TIME_COLUMNS.get(database.strip().lower(), ""),
    )
    ordered = [candidate for candidate in preferred if candidate in columns]
    ordered.extend(column for column in columns if column in TIME_COLUMNS)
    return tuple(dict.fromkeys(value for value in ordered if value))


def _authorized_native_value_columns(
    columns: Sequence[str], declared_concepts: Sequence[str]
) -> set[str]:
    """Bind native physical columns to manifest-selected concept authority."""

    values = [
        column
        for column in columns
        if column not in IDENTIFIER_COLUMNS and column not in TIME_COLUMNS
    ]
    authorized: set[str] = set()
    for concept in declared_concepts:
        if concept in values:
            authorized.add(concept)
            authorized.update(
                f"{concept}{suffix}"
                for suffix in _AUTHORIZED_COMPANION_SUFFIXES
                if f"{concept}{suffix}" in values
            )
            continue
        candidates = [value for value in values if value.startswith(concept + "_")]
        if len(candidates) == 1:
            authorized.add(candidates[0])
        elif len(candidates) > 1:
            raise ExportPackageError(
                f"manifest concept {concept!r} has ambiguous physical outputs: "
                f"{sorted(candidates)!r}",
                code="manifest_concept_physical_mapping_ambiguous",
            )
    return authorized


def open_export_package(export_dir: Union[str, Path]) -> ExportPackage:
    """Parse and verify one native-first EasyICU export package."""

    root = Path(export_dir).expanduser()
    if not root.exists() or not root.is_dir():
        raise FileNotFoundError(f"EasyICU export directory not found: {root}")
    manifest_path, manifest_kind = _select_manifest(root)

    manifest, manifest_bytes = _read_json(manifest_path, label="export manifest")
    database = str(manifest.get("database") or "").strip()
    raw_manifest_format = (
        manifest.get("format")
        if manifest_kind == "native"
        else manifest.get("export_format")
    )
    manifest_format = _normalize_format(raw_manifest_format)
    if raw_manifest_format not in (None, "") and manifest_format is None:
        raise ExportPackageError(
            f"manifest has an unsupported format: {raw_manifest_format!r}",
            code="manifest_format_invalid",
        )
    if manifest_kind == "native" and manifest_format is None:
        raise ExportPackageError(
            "native manifest has an unsupported or missing format",
            code="manifest_format_invalid",
        )

    physical_files: list[ExportPhysicalFile] = []
    concept_index: dict[str, Mapping[str, object]] = {}
    entries = _manifest_file_entries(manifest, manifest_kind=manifest_kind)
    if manifest_kind == "native":
        selected_concepts, selected_by_module = _native_selection(manifest, entries)
    else:
        raw_selected = manifest.get("selected_concepts")
        selected_concepts = (
            {str(value) for value in raw_selected if str(value)}
            if isinstance(raw_selected, list)
            else set()
        )
        selected_by_module = MappingProxyType({})
    resolved_root = root.resolve(strict=True)
    seen_paths: set[str] = set()
    detected_formats: set[str] = set()
    files_by_module: dict[str, set[str]] = {}
    for entry in entries:
        path = _safe_manifest_file(root, entry.get("file"), label="data file")
        relative_path = path.relative_to(resolved_root).as_posix()
        if relative_path in seen_paths:
            raise ExportPackageError(
                f"manifest lists the same file more than once: {relative_path!r}",
                code="manifest_duplicate_file",
                member=relative_path,
            )
        seen_paths.add(relative_path)
        file_format = _FORMAT_BY_SUFFIX.get(path.suffix.lower())
        if file_format is None:
            raise ExportPackageError(
                f"unsupported export file extension: {path.name}",
                code="manifest_format_invalid",
            )
        detected_formats.add(file_format)
        if manifest_format is not None and file_format != manifest_format:
            raise ExportPackageError(
                f"manifest format {manifest_format!r} conflicts with {path.name!r}",
                code="manifest_format_mismatch",
            )
        columns, physical_rows, excel_sheet, identity = _schema_for_file(
            path, file_format
        )
        declared_rows_raw = entry.get("rows")
        if declared_rows_raw is not None and (
            not isinstance(declared_rows_raw, int)
            or isinstance(declared_rows_raw, bool)
        ):
            raise ExportPackageError(
                f"invalid row count for {path.name!r}",
                code="manifest_row_count_invalid",
            )
        declared_rows = declared_rows_raw
        if declared_rows is not None and declared_rows < 0:
            raise ExportPackageError(
                f"negative row count for {path.name!r}",
                code="manifest_row_count_invalid",
            )
        if declared_rows is not None and declared_rows != physical_rows:
            raise ExportPackageError(
                f"row count mismatch for {path.name!r}",
                code="manifest_row_count_mismatch",
            )
        concepts = _declared_concepts(entry)
        module = str(entry.get("module") or entry.get("group") or "")
        files_by_module.setdefault(module, set()).add(relative_path)
        id_column = _identity_column(columns, database, path=path)
        time_columns = _time_columns(columns, database)
        time_column = time_columns[0] if time_columns else None
        physical_file = ExportPhysicalFile(
            path=path,
            relative_path=relative_path,
            module=module,
            file_format=file_format,
            rows=physical_rows,
            columns=columns,
            declared_concepts=concepts,
            id_column=id_column,
            time_column=time_column,
            time_columns=time_columns,
            identity=identity,
            excel_sheet=excel_sheet,
        )
        physical_files.append(physical_file)
        value_columns = (
            _authorized_native_value_columns(columns, concepts)
            if manifest_kind == "native"
            else {
                column
                for column in columns
                if column not in IDENTIFIER_COLUMNS and column not in TIME_COLUMNS
            }
        )
        for column in columns:
            if column not in value_columns:
                continue
            if column in concept_index:
                raise ExportPackageError(
                    f"concept {column!r} appears in more than one manifest-listed file"
                )
            concept_index[column] = MappingProxyType(
                {
                    "file": str(path),
                    "file_name": path.name,
                    "relative_path": relative_path,
                    "rows": physical_file.rows,
                    "columns": list(columns),
                    "format": file_format,
                    "module": physical_file.module,
                    "id_column": id_column,
                    "time_column": time_column,
                    "time_columns": list(time_columns),
                    "sha256": identity.sha256,
                }
            )

    missing_selected_concepts: tuple[str, ...] = ()
    if manifest_kind == "native":
        missing_selected_concepts = tuple(
            sorted(
                concept
                for concept in selected_concepts
                if concept not in concept_index
                and len(
                    [
                        column
                        for column in concept_index
                        if column.startswith(concept + "_")
                    ]
                )
                != 1
            )
        )

    if len(detected_formats) != 1:
        raise ExportPackageError(
            "manifest data files must use one export format",
            code="manifest_format_mismatch",
        )
    export_format = next(iter(detected_formats))

    feature_definitions, feature_sha = _load_feature_definitions(
        root=root,
        manifest=manifest,
        selected_concepts=selected_concepts,
        selected_by_module=selected_by_module,
        files_by_module=files_by_module,
        authorized_files=seen_paths,
        database=database,
    )
    authority_payload = {
        "manifest_kind": manifest_kind,
        "manifest_sha256": hashlib.sha256(manifest_bytes).hexdigest(),
        "database": database,
        "export_format": export_format,
        "feature_definitions_sha256": feature_sha,
        "missing_selected_concepts": list(missing_selected_concepts),
        "files": [
            {
                "relative_path": item.relative_path,
                "sha256": item.identity.sha256,
                "rows": item.rows,
                "columns": list(item.columns),
                "id_column": item.id_column,
                "time_column": item.time_column,
                "time_columns": list(item.time_columns),
            }
            for item in physical_files
        ],
    }
    authority_sha256 = hashlib.sha256(
        json.dumps(authority_payload, sort_keys=True, separators=(",", ":")).encode(
            "utf-8"
        )
    ).hexdigest()
    return ExportPackage(
        root=root.resolve(strict=True),
        manifest_path=manifest_path.resolve(strict=True),
        manifest_kind=manifest_kind,
        manifest_sha256=hashlib.sha256(manifest_bytes).hexdigest(),
        authority_sha256=authority_sha256,
        database=database,
        export_format=export_format,
        files=tuple(physical_files),
        concept_index=MappingProxyType(concept_index),
        missing_selected_concepts=missing_selected_concepts,
        feature_definitions=feature_definitions,
        feature_definitions_sha256=feature_sha,
    )


def verify_export_package(package: ExportPackage) -> None:
    """Re-open ``package`` and require the same content-bound authority."""

    current = open_export_package(package.root)
    if current.authority_sha256 != package.authority_sha256:
        raise ExportPackageError(
            "export package authority changed during materialization",
            code="export_package_authority_changed",
            manifest_path=package.manifest_path,
        )


def index_export_package(export_dir: Union[str, Path]) -> Dict[str, Dict[str, object]]:
    """Return a manifest-authoritative concept index for an export package."""

    return open_export_package(export_dir).index_dict()


def resolve_exported_concept(
    index: Mapping[str, Mapping[str, object]], concept: str
) -> Optional[str]:
    """Resolve exact or one unambiguous ``<concept>_<suffix>`` column."""

    if concept in index:
        return concept
    candidates = sorted(value for value in index if value.startswith(concept + "_"))
    return candidates[0] if len(candidates) == 1 else None


def require_canonical_time_projection(package: ExportPackage, concept: str) -> None:
    """Reject raw database-native time before an hours-based consumer runs."""

    resolved = resolve_exported_concept(package.concept_index, concept)
    if resolved is None:
        return
    info = package.concept_index[resolved]
    raw_time_columns = tuple(
        str(value)
        for value in (info.get("time_columns") or [info.get("time_column")])
        if value not in {None, ""}
    )
    unprojected = [
        value for value in raw_time_columns if value not in {"charttime", "time"}
    ]
    if unprojected or len(raw_time_columns) > 1:
        raise ExportPackageError(
            "export concept uses unprojected or ambiguous time columns "
            f"{raw_time_columns!r}; typed time-unit/origin projection is required "
            "before hours-based materialization",
            code="export_time_projection_required",
            manifest_path=package.manifest_path,
            member=str(info.get("relative_path") or ""),
        )


def _read_projected(
    physical_file: ExportPhysicalFile, columns: Sequence[str]
) -> pd.DataFrame:
    try:
        with _verified_file_snapshot(
            physical_file.path,
            expected=physical_file.identity,
            member=physical_file.relative_path,
        ) as (snapshot, _identity):
            if physical_file.file_format == "parquet":
                frame = pd.read_parquet(
                    snapshot, engine="pyarrow", columns=list(columns)
                )
            elif physical_file.file_format == "csv":
                frame = pd.read_csv(snapshot, usecols=list(columns))
            elif physical_file.file_format == "excel":
                frame = pd.read_excel(
                    snapshot,
                    usecols=list(columns),
                    sheet_name=physical_file.excel_sheet,
                    engine="openpyxl",
                )
            else:  # pragma: no cover - verified during package open
                raise ExportPackageError(
                    f"unsupported export format: {physical_file.file_format}"
                )
    except ExportPackageError:
        raise
    except Exception as exc:
        raise ExportPackageError(
            f"cannot read manifest-listed export file: {physical_file.path}"
        ) from exc
    if physical_file.rows is not None and len(frame) != physical_file.rows:
        raise ExportPackageError(
            f"row count changed after manifest verification: {physical_file.relative_path}"
        )
    return frame


def read_exported_concept(
    export_dir: Union[str, Path, ExportPackage],
    concept: str,
    *,
    extra_columns: Optional[Sequence[str]] = None,
) -> pd.DataFrame:
    """Read one concept from its unique manifest-listed physical file."""

    package = (
        export_dir
        if isinstance(export_dir, ExportPackage)
        else open_export_package(export_dir)
    )
    index = package.concept_index
    resolved = resolve_exported_concept(index, concept)
    if resolved is None:
        raise KeyError(
            f"Concept {concept!r} is not present in {package.root}. "
            f"Available concepts include: {sorted(index)[:20]}"
        )
    info = index[resolved]
    path = Path(str(info["file"]))
    physical_file = next(item for item in package.files if item.path == path)
    authorized = {
        value
        for value, candidate in package.concept_index.items()
        if Path(str(candidate["file"])) == path
    }
    columns = [
        value
        for value in [
            physical_file.id_column,
            *physical_file.time_columns,
            resolved,
            *(extra_columns or []),
        ]
        if value is not None
        if value in {physical_file.id_column, *physical_file.time_columns}
        or value in authorized
    ]
    frame = _read_projected(physical_file, columns)
    rename: dict[str, str] = {}
    if physical_file.id_column != ID_COL:
        rename[physical_file.id_column] = ID_COL
    if physical_file.time_column == "time" and "charttime" not in frame.columns:
        rename["time"] = "charttime"
    if resolved != concept:
        rename[resolved] = concept
    if rename:
        frame = frame.rename(columns=rename)
    return frame


__all__ = [
    "ExportPackage",
    "ExportPackageError",
    "index_export_package",
    "is_export_package",
    "open_export_package",
    "read_exported_concept",
    "require_canonical_time_projection",
    "resolve_exported_concept",
    "verify_export_package",
]
